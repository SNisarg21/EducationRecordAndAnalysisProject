from openpyxl import *
from tkinter import *

wb = load_workbook("D:\\Python_Projects\\PPLProject\\PPLProject.xlsx")
sheet = wb.active


# Function to set focus (cursor) on field box


def ffocus1(event):
    AB_field.focus_set()


def ffocus2(event):
    BB_field.focus_set()


def ffocus3(event):
    BC_field.focus_set()


def ffocus4(event):
    CC_field.focus_set()


def ffocus5(event):
    CD_field.focus_set()


def ffocus6(event):
    DD_field.focus_set()

# Function for clearing the contents of text entry boxes


def cclear():
    AA_field.delete(0, END)
    AB_field.delete(0, END)
    BB_field.delete(0, END)
    BC_field.delete(0, END)
    CC_field.delete(0, END)
    CD_field.delete(0, END)
    DD_field.delete(0, END)

# Function to take data from GUI window and write to an Excel file


lst = []

def Grades():
    if AA_field.get() == "" and AB_field.get() == "" and BB_field.get() == "" and BC_field.get() == "" and CC_field.get() == "" and CD_field.get() == "" and DD_field.get() == "":
        print("empty input")

    else:
        lst.append(int(AA_field.get()))
        lst.append(int(AB_field.get()))
        lst.append(int(BB_field.get()))
        lst.append(int(BC_field.get()))
        lst.append(int(CC_field.get()))
        lst.append(int(CD_field.get()))
        lst.append(int(DD_field.get()))
        # set focus on the name_field box
        AA_field.focus_set()

        # call the clear() function
        cclear()

        return lst


def excel():
    # resize the width of columns in excel spreadsheet

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 10
    sheet.column_dimensions['H'].width = 10
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['K'].width = 25
    sheet.column_dimensions['L'].width = 40

    # write given data to an excel spreadsheet at particular location

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Course"
    sheet.cell(row=1, column=3).value = "Semester"
    sheet.cell(row=1, column=4).value = "MIS Number"
    sheet.cell(row=1, column=5).value = "T1"
    sheet.cell(row=1, column=6).value = "T2"
    sheet.cell(row=1, column=7).value = "ESE"
    sheet.cell(row=1, column=8).value = "TOTAL"
    sheet.cell(row=1, column=9).value = "Grade Point"
    sheet.cell(row=1, column=10).value = "Grade"
    sheet.cell(row=1, column=11).value = "Contact Number"
    sheet.cell(row=1, column=12).value = "Email id"

# Function to set focus (cursor) on field box


def focus1(event):
    course_field.focus_set()


def focus2(event):
    sem_field.focus_set()


def focus3(event):
    MIS_no_field.focus_set()


def focus4(event):
    T1_field.focus_set()


def focus5(event):
    T2_field.focus_set()


def focus6(event):
    ESE_field.focus_set()


def focus7(event):
    contact_no_field.focus_set()


def focus8(event):
    email_id_field.focus_set()

# Function for clearing the contents of text entry boxes


def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    MIS_no_field.delete(0, END)
    T1_field.delete(0, END)
    T2_field.delete(0, END)
    ESE_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)

# Function to take data from GUI window and write to an Excel file


def insert():
    if name_field.get() == "" and course_field.get() == "" and sem_field.get() == "" and MIS_no_field.get() == "" and T1_field.get() == "" and T2_field.get() == "" and ESE_field.get() == "" and contact_no_field.get() == "" and email_id_field.get() == "":
        print("empty input")

    else:
        current_row = sheet.max_row

        # get method returns current text as string which we write into Excel spreadsheet at particular location

        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = int(MIS_no_field.get())
        sheet.cell(row=current_row + 1, column=5).value = float(T1_field.get())
        sheet.cell(row=current_row + 1, column=6).value = float(T2_field.get())
        sheet.cell(row=current_row + 1, column=7).value = float(ESE_field.get())

        Total_marks = sheet.cell(row=current_row + 1, column=5).value + sheet.cell(row=current_row + 1, column=6).value + sheet.cell(row=current_row + 1, column=7).value
        sheet.cell(row=current_row + 1, column=8).value = Total_marks

        if Total_marks >= lst[0]:
            sheet.cell(row=current_row + 1, column=9).value = 10
            sheet.cell(row=current_row + 1, column=10).value = "AA"
        elif lst[1] <= Total_marks < lst[0]:
            sheet.cell(row=current_row + 1, column=9).value = 9
            sheet.cell(row=current_row + 1, column=10).value = "AB"
        elif lst[2] <= Total_marks < lst[1]:
            sheet.cell(row=current_row + 1, column=9).value = 8
            sheet.cell(row=current_row + 1, column=10).value = "BB"
        elif lst[3] <= Total_marks < lst[2]:
            sheet.cell(row=current_row + 1, column=9).value = 7
            sheet.cell(row=current_row + 1, column=10).value = "BC"
        elif lst[4] <= Total_marks < lst[3]:
            sheet.cell(row=current_row + 1, column=9).value = 6
            sheet.cell(row=current_row + 1, column=10).value = "CC"
        elif lst[5] <= Total_marks < lst[4]:
            sheet.cell(row=current_row + 1, column=9).value = 5
            sheet.cell(row=current_row + 1, column=10).value = "CD"
        elif lst[6] <= Total_marks < lst[5]:
            sheet.cell(row=current_row + 1, column=9).value = 9
            sheet.cell(row=current_row + 1, column=10).value = "DD"
        else:
            sheet.cell(row=current_row + 1, column=9).value = 'FAIL'
            sheet.cell(row=current_row + 1, column=10).value = 'FAIL'

        sheet.cell(row=current_row + 1, column=11).value = int(contact_no_field.get())
        sheet.cell(row=current_row + 1, column=12).value = (email_id_field.get())

        # save the file
        wb.save("D:\\Python_Projects\\PPLProject\\PPLProject.xlsx")

        # set focus on the name_field box
        name_field.focus_set()

        # call the clear() function
        clear()


# Driver code

if __name__ == "__main__":


    # create a GUI window
    root1 = Tk()

    # set the background colour of GUI window
    root1.configure(background='light blue')

    # set the title of GUI window
    root1.title("Student Information")

    # set the configuration of GUI window
    root1.geometry("500x250")


    # creating labels

    heading = Label(root1, text="Enter Grade Details of the course!", bg="light blue")
    AA = Label(root1, text="AA", bg="light blue")
    AB = Label(root1, text="AB", bg="light blue")
    BB = Label(root1, text="BB", bg="light blue")
    BC = Label(root1, text="BC", bg="light blue")
    CC = Label(root1, text="CC", bg="light blue")
    CD = Label(root1, text="CD", bg="light blue")
    DD = Label(root1, text="DD", bg="light blue")

    # grid method is used for placing the widgets at respective positions in table like structure .

    heading.grid(row=0, column=1)
    AA.grid(row=1, column=0)
    AB.grid(row=2, column=0)
    BB.grid(row=3, column=0)
    BC.grid(row=4, column=0)
    CC.grid(row=5, column=0)
    CD.grid(row=6, column=0)
    DD.grid(row=7, column=0)

    # create a text entry box for typing the details

    AA_field = Entry(root1)
    AB_field = Entry(root1)
    BB_field = Entry(root1)
    BC_field = Entry(root1)
    CC_field = Entry(root1)
    CD_field = Entry(root1)
    DD_field = Entry(root1)

    # bind method of widget is used for the binding the function with the events whenever the enter key is pressed then call the next focus function

    AA_field.bind("<Return>", ffocus1)
    AB_field.bind("<Return>", ffocus2)
    BB_field.bind("<Return>", ffocus3)
    BC_field.bind("<Return>", ffocus4)
    CC_field.bind("<Return>", ffocus5)
    CD_field.bind("<Return>", ffocus6)
    DD_field.bind("<Return>")

    # grid method is used for placing the widgets at respective positions in table like structure .
    AA_field.grid(row=1, column=1, ipadx="100")
    AB_field.grid(row=2, column=1, ipadx="100")
    BB_field.grid(row=3, column=1, ipadx="100")
    BC_field.grid(row=4, column=1, ipadx="100")
    CC_field.grid(row=5, column=1, ipadx="100")
    CD_field.grid(row=6, column=1, ipadx="100")
    DD_field.grid(row=7, column=1, ipadx="100")

    # create a Submit Button and place into the root window
    submit = Button(root1, text="Submit", fg="Black", bg="Red", command=Grades)
    submit.grid(row=8, column=1)

    # start the GUI
    root1.mainloop()


    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='light blue')

    # set the title of GUI window
    root.title("Student Information")

    # set the configuration of GUI window
    root.geometry("500x250")

    excel()

    # creating labels

    heading = Label(root, text="Enter Details", bg="light blue")
    name = Label(root, text="Name", bg="light blue")
    course = Label(root, text="Course", bg="light blue")
    sem = Label(root, text="Semester", bg="light blue")
    MIS_no = Label(root, text="MIS No.", bg="light blue")
    T1 = Label(root, text="T1", bg="light blue")
    T2 = Label(root, text="T2", bg="light blue")
    ESE = Label(root, text="ESE", bg="light blue")
    contact_no = Label(root, text="Contact No.", bg="light blue")
    email_id = Label(root, text="Email id", bg="light blue")

    # grid method is used for placing the widgets at respective positions in table like structure .

    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    course.grid(row=2, column=0)
    sem.grid(row=3, column=0)
    MIS_no.grid(row=4, column=0)
    T1.grid(row=5, column=0)
    T2.grid(row=6, column=0)
    ESE.grid(row=7, column=0)
    contact_no.grid(row=8, column=0)
    email_id.grid(row=9, column=0)

    # create a text entry box for typing the details

    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    MIS_no_field = Entry(root)
    T1_field = Entry(root)
    T2_field = Entry(root)
    ESE_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)

    # bind method of widget is used for the binding the function with the events whenever the enter key is pressed then call the next focus function

    name_field.bind("<Return>", focus1)
    course_field.bind("<Return>", focus2)
    sem_field.bind("<Return>", focus3)
    MIS_no_field.bind("<Return>", focus4)
    T1_field.bind("<Return>", focus5)
    T2_field.bind("<Return>", focus6)
    ESE_field.bind("<Return>", focus7)
    contact_no_field.bind("<Return>", focus8)
    email_id_field.bind("<Return>")

    # grid method is used for placing the widgets at respective positions in table like structure .
    name_field.grid(row=1, column=1, ipadx="100")
    course_field.grid(row=2, column=1, ipadx="100")
    sem_field.grid(row=3, column=1, ipadx="100")
    MIS_no_field.grid(row=4, column=1, ipadx="100")
    T1_field.grid(row=5, column=1, ipadx="100")
    T2_field.grid(row=6, column=1, ipadx="100")
    ESE_field.grid(row=7, column=1, ipadx="100")
    contact_no_field.grid(row=8, column=1, ipadx="100")
    email_id_field.grid(row=9, column=1, ipadx="100")

    # call excel function
    excel()

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black", bg="Red", command=insert)
    submit.grid(row=10, column=1)

    # start the GUI
    root.mainloop()
