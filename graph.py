from openpyxl import *
from openpyxl.chart import BarChart, Reference
import main

wb = load_workbook("D:\\Python_Projects\\PPLProject\\PPLProject.xlsx")
sheetC = wb.create_sheet('Analysis', 1)

countAA = 0
countAB = 0
countBB = 0
countBC = 0
countCC = 0
countCD = 0
countDD = 0

for i in range(2, main.sheet.max_row+1):
    if main.sheet.cell(row= i, column=9).value == 10:
        countAA += 1
    elif main.sheet.cell(row= i, column=9).value == 9:
        countAB += 1
    elif main.sheet.cell(row= i, column=9).value == 8:
        countBB += 1
    elif main.sheet.cell(row= i, column=9).value == 7:
        countBC += 1
    elif main.sheet.cell(row= i, column=9).value == 6:
        countCC += 1
    elif main.sheet.cell(row= i, column=9).value == 10:
        countCD += 1
    else:
        countDD += 1

sheetC.cell(row=2,column=2).value = countAA
sheetC.cell(row=3,column=2).value = countAB
sheetC.cell(row=4,column=2).value = countBB
sheetC.cell(row=5,column=2).value = countBC
sheetC.cell(row=6,column=2).value = countCC
sheetC.cell(row=7,column=2).value = countCD
sheetC.cell(row=8,column=2).value = countDD

sheetC.cell(row=2,column=1).value = 'AA'
sheetC.cell(row=3,column=1).value = 'AB'
sheetC.cell(row=4,column=1).value = 'BB'
sheetC.cell(row=5,column=1).value = 'BC'
sheetC.cell(row=6,column=1).value = 'CC'
sheetC.cell(row=7,column=1).value = 'CD'
sheetC.cell(row=8,column=1).value = 'DD'

data = Reference(sheetC, min_col=1, min_row=1, max_col=2, max_row=8)
titles = Reference(sheetC, min_col=1, min_row=2, max_row=8)
chart = BarChart()
chart.title = "Grade Analysis"
chart.add_data(data=data, titles_from_data=True)
chart.set_categories(titles)
sheetC.add_chart(chart, 'A10')

wb.save("D:\\Python_Projects\\PPLProject\\PPLProject.xlsx")
